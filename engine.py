# -*- coding: utf-8 -*-
"""
Stock-turn posting engine.
Parses the daily 1C warehouse movement export (with manual operation keys)
and posts signed movements into the accounting workbook (ნაშთები + ჟურნალი).
Matching key: article (არტიკული), normalized to text. 1C internal code (კოდი)
is kept as a reference column only.
Kept UI-free so it can be unit-tested; app.py provides the Streamlit front-end.
"""
import io
import re
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------- constants

KEY_MAP = {
    "მ": ("მიღება", 1),
    "დ": ("დაბრუნება", 1),
    "გ": ("გაცემა", -1),
    "ჩ": ("ჩამოწერა", -1),
}

BAL_SHEET = "ნაშთები"
LED_SHEET = "ჟურნალი"
DAILY_SHEET = "დღიური ნაშთები"
INSTR_SHEET = "ინსტრუქცია"

BAL_COLS = ["არტიკული", "ნომენკლატურა", "კოდი", "ძირითადი შტრიხ-კოდი",
            "საწყისი ნაშთი", "მიმდინარე ნაშთი"]
LED_COLS = ["თარიღი", "დოკუმენტი", "ოპერაცია", "ოპერაციის ტიპი", "არტიკული",
            "ნომენკლატურა", "რაოდენობა", "ცვლილება (+/-)", "ატვირთვის დრო",
            "წყარო ფაილი"]

DOC_RE = re.compile(r"([A-ZА-Я]?\d{7,})\s+თარიღით\s+(\d{2}\.\d{2}\.\d{4})(?:\s+(\d{2}:\d{2}:\d{2}))?")


def normalize_article(v) -> str:
    """Article as clean text: 33205.0 -> '33205'; alphanumeric kept as-is."""
    if pd.isna(v):
        return ""
    s = str(v).strip()
    try:
        f = float(s)
        if f == int(f):
            return str(int(f))
    except (ValueError, OverflowError):
        pass
    return s


# ---------------------------------------------------------------- daily file

def parse_daily_file(file_bytes: bytes, filename: str):
    """Return dict with transactions, doc column info, warnings, errors."""
    engine = "xlrd" if filename.lower().endswith(".xls") else "openpyxl"
    raw = pd.read_excel(io.BytesIO(file_bytes), header=None, engine=engine)

    errors, warnings = [], []

    # --- locate header row: contains both 'ნომენკლატურა' and an article column
    header_row = None
    for i in range(min(15, len(raw))):
        vals = [str(v).strip() for v in raw.iloc[i] if pd.notna(v)]
        if "ნომენკლატურა" in vals and any(v.startswith("არტიკული") for v in vals):
            header_row = i
            break
    if header_row is None:
        return {"errors": ["ვერ ვიპოვე სათაურის სტრიქონი (ნომენკლატურა / არტიკული). "
                           "დარწმუნდით, რომ ეს არის 1C-ის მოძრაობის რეპორტი."],
                "warnings": [], "transactions": pd.DataFrame(), "docs": []}
    if header_row == 0:
        return {"errors": ["სათაურის ზემოთ სტრიქონი არ არსებობს — ოპერაციის ასოების ჩასაწერი ადგილი ვერ მოიძებნა."],
                "warnings": [], "transactions": pd.DataFrame(), "docs": []}

    hdr = raw.iloc[header_row]
    key_row = raw.iloc[header_row - 1]

    def col_of(name):
        for j, v in hdr.items():
            if pd.notna(v) and str(v).strip() == name:
                return j
        return None

    c_name = col_of("ნომენკლატურა")
    c_num = col_of("№")
    c_total = col_of("სულ")
    c_art = next((j for j, v in hdr.items()
                  if pd.notna(v) and str(v).strip().startswith("არტიკული")), None)
    if c_art is None:
        return {"errors": ["ფაილში ვერ მოიძებნა სვეტი \"არტიკული\" — დამთხვევა შეუძლებელია."],
                "warnings": [], "transactions": pd.DataFrame(), "docs": []}

    # --- document columns: header contains 'თარიღით'
    docs = []
    for j, v in hdr.items():
        if pd.isna(v) or "თარიღით" not in str(v):
            continue
        text = str(v).strip()
        m = DOC_RE.search(text)
        doc_no, doc_date, doc_time = (m.group(1), m.group(2), m.group(3) or "") if m else (text, "", "")
        doc_type = text.split(m.group(1))[0].strip() if m else text
        key = key_row[j]
        key = str(key).strip() if pd.notna(key) else ""
        docs.append({"col": j, "header": text, "doc_type": doc_type,
                     "doc_no": doc_no, "date": doc_date, "time": doc_time, "key": key})

    if not docs:
        errors.append("ფაილში ოპერაციის (დოკუმენტის) სვეტები ვერ მოიძებნა.")

    # --- key validation
    for d in docs:
        if d["key"] == "":
            errors.append(f'სვეტს არ აქვს ოპერაციის ასო: "{d["doc_type"]} {d["doc_no"]}" '
                          f'(სვეტი {get_column_letter(d["col"] + 1)}). ჩაწერეთ მ / დ / გ / ჩ სათაურის ზემოთ.')
        elif d["key"] not in KEY_MAP:
            errors.append(f'უცნობი ოპერაციის ასო "{d["key"]}" სვეტზე "{d["doc_type"]} {d["doc_no"]}". '
                          f"დასაშვებია მხოლოდ: მ, დ, გ, ჩ.")

    # --- soft consistency warnings (doc type vs key)
    for d in docs:
        if "ჩამოწერა" in d["doc_type"] and d["key"] not in ("ჩ", ""):
            warnings.append(f'დოკუმენტი "{d["doc_no"]}" 1C-ში ჩამოწერაა, მაგრამ მონიშნულია "{d["key"]}" — გადაამოწმეთ.')
        if "შემოსვლა" in d["doc_type"] and d["key"] in ("გ", "ჩ"):
            warnings.append(f'დოკუმენტი "{d["doc_no"]}" 1C-ში შემოსვლაა, მაგრამ მონიშნულია გამავალად ("{d["key"]}") — გადაამოწმეთ.')

    if errors:
        return {"errors": errors, "warnings": warnings,
                "transactions": pd.DataFrame(), "docs": docs}

    # --- data rows: below header, article present, not the total row
    body = raw.iloc[header_row + 1:].copy()
    num_col = c_num if c_num is not None else raw.columns[0]
    body = body[body[num_col].astype(str).str.strip() != "სულ"]
    body = body[body[c_art].notna()]
    body["_art"] = body[c_art].map(normalize_article)
    body = body[body["_art"] != ""]

    dup_arts = body["_art"].duplicated()
    if dup_arts.any():
        warnings.append(f"ყოველდღიურ ფაილში {int(dup_arts.sum())} დუბლირებული არტიკულია — რაოდენობები დაჯამდება.")

    tx = []
    for _, row in body.iterrows():
        art = row["_art"]
        name = str(row[c_name]).strip() if pd.notna(row[c_name]) else ""
        for d in docs:
            q = pd.to_numeric(row[d["col"]], errors="coerce")
            if pd.isna(q) or q == 0:
                if pd.notna(row[d["col"]]) and str(row[d["col"]]).strip() not in ("", " "):
                    if pd.isna(q):
                        warnings.append(f'არაციფრული მნიშვნელობა "{row[d["col"]]}" (არტიკული {art}, დოკ. {d["doc_no"]}) — გამოტოვებულია.')
                continue
            op_name, sign = KEY_MAP[d["key"]]
            tx.append({"თარიღი": d["date"], "დოკუმენტი": d["doc_no"],
                       "ოპერაცია": f'{d["key"]} — {op_name}', "ოპერაციის ტიპი": d["doc_type"],
                       "არტიკული": art, "ნომენკლატურა": name,
                       "რაოდენობა": float(q), "ცვლილება (+/-)": sign * float(q)})
    tx = pd.DataFrame(tx)

    # --- integrity check vs the 1C 'სულ' column (it sums absolute movements)
    if c_total is not None and not tx.empty:
        file_total = pd.to_numeric(body[c_total], errors="coerce").sum()
        parsed_total = tx["რაოდენობა"].sum()
        if abs(file_total - parsed_total) > 0.01:
            warnings.append(f"საკონტროლო ჯამი არ ემთხვევა: 1C 'სულ' = {file_total:,.0f}, "
                            f"წაკითხული = {parsed_total:,.0f}. გადაამოწმეთ ფაილი.")

    return {"errors": [], "warnings": warnings, "transactions": tx, "docs": docs}


# ---------------------------------------------------------------- stock file

def load_stock_file(file_bytes: bytes):
    """Read balances (openings) and ledger from the accounting workbook."""
    bio = io.BytesIO(file_bytes)
    bal = pd.read_excel(bio, sheet_name=BAL_SHEET, engine="openpyxl")
    bio.seek(0)
    led = pd.read_excel(bio, sheet_name=LED_SHEET, engine="openpyxl")

    problems = []
    for c in BAL_COLS[:5]:
        if c not in bal.columns:
            problems.append(f'ნაშთების ფურცელს აკლია სვეტი "{c}". (ძველი ვერსიის ფაილი? საჭიროა ახალი შაბლონი, სადაც არტიკული პირველი სვეტია.)')
    for c in LED_COLS[:8]:
        if c not in led.columns:
            problems.append(f'ჟურნალის ფურცელს აკლია სვეტი "{c}".')
    if problems:
        return None, None, problems

    bal = bal[bal["არტიკული"].notna()].copy()
    bal["არტიკული"] = bal["არტიკული"].map(normalize_article)
    bal = bal[bal["არტიკული"] != ""]
    bal["საწყისი ნაშთი"] = pd.to_numeric(bal["საწყისი ნაშთი"], errors="coerce").fillna(0)
    if bal["არტიკული"].duplicated().any():
        problems.append("ნაშთების ფურცელზე დუბლირებული არტიკულებია — გაასწორეთ ფაილი.")
        return None, None, problems

    if not led.empty:
        led["არტიკული"] = led["არტიკული"].map(normalize_article)
        led["ცვლილება (+/-)"] = pd.to_numeric(led["ცვლილება (+/-)"], errors="coerce").fillna(0)
    return bal, led, []


def compute_balances(bal: pd.DataFrame, led: pd.DataFrame) -> pd.DataFrame:
    """Current stock = opening + net of all ledger postings (recomputed, never trusted from file)."""
    out = bal.copy()
    if led is not None and not led.empty:
        net = led.groupby("არტიკული")["ცვლილება (+/-)"].sum()
        out["მიმდინარე ნაშთი"] = out["საწყისი ნაშთი"] + out["არტიკული"].map(net).fillna(0)
    else:
        out["მიმდინარე ნაშთი"] = out["საწყისი ნაშთი"]
    return out


def find_duplicates(led: pd.DataFrame, tx: pd.DataFrame):
    """Doc numbers already posted (protection against posting a day twice)."""
    if led is None or led.empty or tx.empty:
        return []
    posted = set(zip(led["დოკუმენტი"].astype(str), led["თარიღი"].astype(str)))
    dups = sorted({(d, t) for d, t in zip(tx["დოკუმენტი"].astype(str), tx["თარიღი"].astype(str))
                   if (d, t) in posted})
    return dups


def post_and_build(bal, led, tx, source_name: str):
    """Append transactions to ledger, add unknown SKUs, rebuild the workbook. Returns (bytes, info)."""
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    tx = tx.copy()
    tx["ატვირთვის დრო"] = now
    tx["წყარო ფაილი"] = source_name

    # unknown SKUs from the DAILY file -> add to balances with opening 0
    known = set(bal["არტიკული"])
    new_skus = tx.loc[~tx["არტიკული"].isin(known), ["არტიკული", "ნომენკლატურა"]].drop_duplicates("არტიკული")
    for _, r in new_skus.iterrows():
        bal = pd.concat([bal, pd.DataFrame([{
            "არტიკული": r["არტიკული"], "ნომენკლატურა": r["ნომენკლატურა"],
            "კოდი": "", "ძირითადი შტრიხ-კოდი": "", "საწყისი ნაშთი": 0}])],
            ignore_index=True)

    led_new = pd.concat([led, tx[LED_COLS]], ignore_index=True) if led is not None and not led.empty \
        else tx[LED_COLS].copy()

    balances = compute_balances(bal, led_new)
    negative = balances[balances["მიმდინარე ნაშთი"] < 0]

    data = _build_workbook(bal, led_new, set(new_skus["არტიკული"]))
    info = {"new_skus": new_skus, "negative": negative, "balances": balances,
            "posted_rows": len(tx)}
    return data, info


# ---------------------------------------------------------------- 1C compare

def parse_1c_stock_export(file_bytes: bytes, filename: str):
    """Parse the 1C report 'უწყისი — საქონელი საწყობში'.
    Returns (DataFrame[არტიკული, ნომენკლატურა, 1C ნაშთი], errors).
    Group/hierarchy rows (no article) are skipped."""
    engine = "xlrd" if filename.lower().endswith(".xls") else "openpyxl"
    raw = pd.read_excel(io.BytesIO(file_bytes), header=None, engine=engine)

    header_row = None
    for i in range(min(20, len(raw))):
        vals = [str(v).strip() for v in raw.iloc[i] if pd.notna(v)]
        if "ნომენკლატურა" in vals and "საბოლოო ნაშთი" in vals:
            header_row = i
            break
    if header_row is None:
        return None, ["ვერ ვიპოვე 1C ნაშთების რეპორტის სათაური (ნომენკლატურა / საბოლოო ნაშთი). "
                      "ატვირთეთ რეპორტი \"უწყისი — საქონელი საწყობში\"."]

    hdr = raw.iloc[header_row]
    c_name = next(j for j, v in hdr.items() if pd.notna(v) and str(v).strip() == "ნომენკლატურა")
    c_close = next(j for j, v in hdr.items() if pd.notna(v) and str(v).strip() == "საბოლოო ნაშთი")
    c_open = next((j for j, v in hdr.items() if pd.notna(v) and str(v).strip() == "საწყისი ნაშთი"), c_close)
    # unlabeled columns between name and 'საწყისი ნაშთი' are the additional
    # fields (article, then barcode)
    extra = [j for j in hdr.index if c_name < j < c_open]
    if not extra:
        return None, ["რეპორტში ვერ მოიძებნა არტიკულის სვეტი (დამატებითი ველი ნომენკლატურასა და ნაშთებს შორის)."]
    c_art = extra[0]

    body = raw.iloc[header_row + 1:].copy()
    body = body[body[c_art].notna()]
    body["_art"] = body[c_art].map(normalize_article)
    body = body[body["_art"] != ""]
    out = pd.DataFrame({
        "არტიკული": body["_art"],
        "ნომენკლატურა (1C)": body[c_name].astype(str).str.strip(),
        "1C ნაშთი": pd.to_numeric(body[c_close], errors="coerce").fillna(0),
    })
    if out["არტიკული"].duplicated().any():
        d = int(out["არტიკული"].duplicated().sum())
        out = out.groupby("არტიკული", as_index=False).agg(
            {"ნომენკლატურა (1C)": "first", "1C ნაშთი": "sum"})
        return out, [f"1C რეპორტში {d} დუბლირებული არტიკული დაჯამდა."]
    return out, []


def compare_with_1c(balances: pd.DataFrame, export: pd.DataFrame):
    """Compare our current stock vs the 1C export closing stock, on article.
    Only articles present in OUR balances are compared; export-only articles
    are returned separately for information and are never added."""
    ours = balances[["არტიკული", "ნომენკლატურა", "მიმდინარე ნაშთი"]].copy()
    ours = ours.rename(columns={"მიმდინარე ნაშთი": "ჩვენი ნაშთი"})
    cmp = ours.merge(export[["არტიკული", "1C ნაშთი"]], on="არტიკული", how="left")
    cmp["1C-ში არსებობს"] = cmp["არტიკული"].isin(set(export["არტიკული"]))
    cmp["1C ნაშთი"] = cmp["1C ნაშთი"].fillna(0)
    cmp["სხვაობა (ჩვენი − 1C)"] = cmp["ჩვენი ნაშთი"] - cmp["1C ნაშთი"]
    export_only = export[~export["არტიკული"].isin(set(ours["არტიკული"]))].copy()
    return cmp, export_only


def build_comparison_file(cmp: pd.DataFrame, export_only: pd.DataFrame, label: str) -> bytes:
    """Standalone comparison report .xlsx (not written into the accounting file)."""
    wb = Workbook()
    thin = Border(*[Side(style="thin", color="B0B0B0")] * 4)
    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    dif_fill = PatternFill("solid", fgColor="FFC7CE")
    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    cell_font = Font(name="Arial", size=10)

    ws = wb.active
    ws.title = "შედარება"
    ws.cell(1, 1, f"ნაშთების შედარება 1C-სთან — {label}").font = Font(name="Arial", bold=True, size=11)
    cols = ["არტიკული", "ნომენკლატურა", "ჩვენი ნაშთი", "1C ნაშთი", "სხვაობა (ჩვენი − 1C)", "1C-ში არსებობს"]
    for j, h in enumerate(cols, 1):
        c = ws.cell(3, j, h)
        c.font, c.fill, c.border = hdr_font, hdr_fill, thin
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, (_, r) in enumerate(cmp.iterrows()):
        row = i + 4
        vals = [r["არტიკული"], r["ნომენკლატურა"], float(r["ჩვენი ნაშთი"]),
                float(r["1C ნაშთი"]), float(r["სხვაობა (ჩვენი − 1C)"]),
                "კი" if r["1C-ში არსებობს"] else "არა"]
        for j, v in enumerate(vals, 1):
            c = ws.cell(row, j, v)
            c.font, c.border = cell_font, thin
        if abs(r["სხვაობა (ჩვენი − 1C)"]) > 0.001:
            for j in range(1, 7):
                ws.cell(row, j).fill = dif_fill
    for j, w in zip(range(1, 7), [14, 55, 12, 12, 16, 13]):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:F{len(cmp) + 3}"

    ws2 = wb.create_sheet("მხოლოდ 1C-შია")
    for j, h in enumerate(["არტიკული", "ნომენკლატურა (1C)", "1C ნაშთი"], 1):
        c = ws2.cell(1, j, h)
        c.font, c.fill, c.border = hdr_font, hdr_fill, thin
    for i, (_, r) in enumerate(export_only.iterrows()):
        ws2.cell(i + 2, 1, r["არტიკული"]).font = cell_font
        ws2.cell(i + 2, 2, r["ნომენკლატურა (1C)"]).font = cell_font
        ws2.cell(i + 2, 3, float(r["1C ნაშთი"])).font = cell_font
    for j, w in zip(range(1, 4), [24, 55, 12]):
        ws2.column_dimensions[get_column_letter(j)].width = w
    ws2.freeze_panes = "A2"

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ---------------------------------------------------------------- workbook

def _build_daily_sheet(wb, bal, led, hdr_font, hdr_fill, cell_font, thin):
    """SKU × day matrix: per posting day, net movement + closing balance.
    Fully regenerated from the ledger on every posting, so later corrections
    of opening stock are always reflected."""
    ws = wb.create_sheet(DAILY_SHEET)

    if led is None or led.empty:
        days = []
        net = None
    else:
        led = led.copy()
        led["_d"] = pd.to_datetime(led["თარიღი"], format="%d.%m.%Y", errors="coerce")
        _SENTINEL = pd.Timestamp("2200-01-01")
        led.loc[led["_d"].isna(), "_d"] = _SENTINEL  # undated postings go last
        days = sorted(led["_d"].unique())
        net = led.pivot_table(index="არტიკული", columns="_d",
                              values="ცვლილება (+/-)", aggfunc="sum")

    ws.cell(1, 1, "არტიკული"); ws.cell(1, 2, "ნომენკლატურა")
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
    for k, d in enumerate(days):
        c0 = 3 + 2 * k
        label = "უცნობი თარიღი" if pd.Timestamp(d).year == 2200 else pd.Timestamp(d).strftime("%d.%m.%Y")
        ws.cell(1, c0, label)
        ws.merge_cells(start_row=1, start_column=c0, end_row=1, end_column=c0 + 1)
        ws.cell(2, c0, "ცვლილება"); ws.cell(2, c0 + 1, "ნაშთი")
    last_col = 2 + 2 * len(days)
    for r in (1, 2):
        for j in range(1, max(last_col, 2) + 1):
            c = ws.cell(r, j)
            c.font, c.fill, c.border = hdr_font, hdr_fill, thin
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, (_, srow) in enumerate(bal.iterrows()):
        row = i + 3
        art = srow["არტიკული"]
        ws.cell(row, 1, art)
        ws.cell(row, 2, srow["ნომენკლატურა"])
        running = float(srow["საწყისი ნაშთი"])
        for k, d in enumerate(days):
            chg = 0.0
            if net is not None and art in net.index and pd.notna(net.loc[art].get(d)):
                chg = float(net.loc[art][d])
            running += chg
            c0 = 3 + 2 * k
            if chg != 0:
                ws.cell(row, c0, chg)
            ws.cell(row, c0 + 1, running)
        for j in range(1, max(last_col, 2) + 1):
            ws.cell(row, j).font, ws.cell(row, j).border = cell_font, thin
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 55
    for j in range(3, max(last_col, 2) + 1):
        ws.column_dimensions[get_column_letter(j)].width = 11
    ws.freeze_panes = "C3"


def _build_workbook(bal, led, new_sku_arts=None) -> bytes:
    new_sku_arts = new_sku_arts or set()
    wb = Workbook()
    thin = Border(*[Side(style="thin", color="B0B0B0")] * 4)
    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    inp_fill = PatternFill("solid", fgColor="FFFF00")
    new_fill = PatternFill("solid", fgColor="FFD8A8")
    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    cell_font = Font(name="Arial", size=10)

    def header(ws, cols, widths):
        for j, h in enumerate(cols, 1):
            c = ws.cell(1, j, h)
            c.font, c.fill, c.border = hdr_font, hdr_fill, thin
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for j, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(j)].width = w
        ws.freeze_panes = "A2"

    ws = wb.active
    ws.title = BAL_SHEET
    header(ws, BAL_COLS, [14, 55, 9, 18, 14, 15])
    for i, (_, r) in enumerate(bal.iterrows()):
        row = i + 2
        ws.cell(row, 1, str(r["არტიკული"]))  # text — SUMIFS matches text vs text
        ws.cell(row, 2, r["ნომენკლატურა"])
        ws.cell(row, 3, "" if pd.isna(r.get("კოდი")) else str(r.get("კოდი")))
        ws.cell(row, 4, "" if pd.isna(r.get("ძირითადი შტრიხ-კოდი")) else str(r.get("ძირითადი შტრიხ-კოდი")))
        oc = ws.cell(row, 5, float(r["საწყისი ნაშთი"]))
        oc.fill = new_fill if r["არტიკული"] in new_sku_arts else inp_fill
        ws.cell(row, 6, f"=E{row}+SUMIFS({LED_SHEET}!H:H,{LED_SHEET}!E:E,A{row})")
        for j in range(1, 7):
            ws.cell(row, j).font, ws.cell(row, j).border = cell_font, thin
    ws.auto_filter.ref = f"A1:F{len(bal) + 1}"

    ws2 = wb.create_sheet(LED_SHEET)
    header(ws2, LED_COLS, [12, 16, 16, 26, 14, 55, 12, 14, 18, 28])
    for i, (_, r) in enumerate(led.iterrows()):
        row = i + 2
        vals = [r["თარიღი"], r["დოკუმენტი"], r["ოპერაცია"], r["ოპერაციის ტიპი"],
                str(r["არტიკული"]), r["ნომენკლატურა"], float(r["რაოდენობა"]),
                float(r["ცვლილება (+/-)"]), r["ატვირთვის დრო"], r["წყარო ფაილი"]]
        for j, v in enumerate(vals, 1):
            c = ws2.cell(row, j, v)
            c.font, c.border = cell_font, thin
    ws2.auto_filter.ref = f"A1:J{max(len(led), 1) + 1}"

    _build_daily_sheet(wb, bal, led, hdr_font, hdr_fill, cell_font, thin)

    ws3 = wb.create_sheet(INSTR_SHEET)
    lines = [
        ("ინსტრუქცია", True),
        ("", False),
        ("1. ყვითელი უჯრები (\"საწყისი ნაშთი\") ივსება ერთხელ, პირველი გამოყენების წინ.", False),
        ("2. \"მიმდინარე ნაშთი\" ითვლება ავტომატურად: საწყისი ნაშთი + ჟურნალის ჩანაწერები. ხელით ნუ შეცვლით.", False),
        ("3. ჟურნალსა და \"დღიური ნაშთები\" ფურცელს ავსებს აპლიკაცია — ხელით ჩანაწერები არ გააკეთოთ.", False),
        ("4. დამთხვევა ხდება არტიკულით. არტიკული ინახება როგორც ტექსტი — ნუ გადააფორმატებთ რიცხვად.", False),
        ("5. ყოველდღიურ ფაილში ოპერაციის სვეტის სათაურის ზემოთ ჩაწერეთ ერთი ასო:", False),
        ("      მ = მიღება (+)      დ = დაბრუნება (+)      გ = გაცემა (−)      ჩ = ჩამოწერა (−)", False),
        ("6. ნარინჯისფერი უჯრა ნიშნავს ახალ SKU-ს, რომელიც ავტომატურად დაემატა — შეავსეთ კოდი/შტრიხ-კოდი საჭიროებისას.", False),
        ("7. აპლიკაციიდან გადმოწერილი ფაილით ჩაანაცვლეთ ძველი ვერსია.", False),
    ]
    for i, (t, b) in enumerate(lines, 1):
        c = ws3.cell(i, 1, t)
        c.font = Font(name="Arial", size=11 if b else 10, bold=b)
    ws3.column_dimensions["A"].width = 120

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
